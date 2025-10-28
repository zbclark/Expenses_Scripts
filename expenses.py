import json  
import datetime  
import csv
import sys
import re
import os
import subprocess
from collections import defaultdict

# Try to import openpyxl for Excel functionality
try:
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  Note: openpyxl not available. Install with 'pip install openpyxl' for Excel features.")  
  
class ExpenseTracker:  
    def __init__(self, filename="expenses.json"):  
        self.filename = filename  
        self.expenses = self.load_expenses()  
      
    def load_expenses(self):  
        """Load expenses from file, create empty list if file doesn't exist"""  
        try:  
            with open(self.filename, 'r') as file:  
                return json.load(file)  
        except FileNotFoundError:  
            return []  
      
    def save_expenses(self):  
        """Save expenses to file"""  
        with open(self.filename, 'w') as file:  
            json.dump(self.expenses, file, indent=2)  
      
    def add_expense(self, amount, category, description=""):  
        """Add a new expense"""  
        expense = {  
            "date": datetime.date.today().isoformat(),  
            "amount": float(amount),  
            "category": category.lower(),  
            "description": description  
        }  
        self.expenses.append(expense)  
        self.save_expenses()  
        print(f"Added expense: ${amount} for {category}")  

    def load_from_csv(self, csv_filename):
        """Load expenses from a CSV file with enhanced bank statement support"""
        # Clear existing expenses to avoid duplicates
        self.expenses = []
        loaded_count = 0
        try:
            # First, clean the CSV file to handle BOM and empty lines
            cleaned_filename = self._clean_csv_file(csv_filename)
            
            with open(cleaned_filename, 'r', newline='', encoding='utf-8') as csvfile:
                # Try to detect the delimiter
                sample = csvfile.read(1024)
                csvfile.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                reader = csv.DictReader(csvfile, delimiter=delimiter)
                
                # Print detected columns for debugging
                print(f"Detected CSV columns: {reader.fieldnames}")
                
                for row_num, row in enumerate(reader, 1):
                    try:
                        # Try different common column name variations
                        amount = self._get_csv_value(row, ['amount', 'cost', 'price', 'total', 'value', 'amount ($)'])
                        category = self._get_csv_value(row, ['category', 'class', 'group'])  # Remove 'type' for bank statements
                        description = self._get_csv_value(row, ['action', 'description', 'desc', 'note', 'memo', 'details'])  # Prioritize 'action' for bank statements
                        date = self._get_csv_value(row, ['date', 'timestamp', 'created', 'when', 'run date'])
                        notes = self._get_csv_value(row, ['notes'])  # Only check for notes if it exists (from our own exports)
                        
                        if amount is not None:
                            # Clean up amount (remove $ and commas, handle negatives)
                            amount_str = str(amount).replace('$', '').replace(',', '')
                            try:
                                amount_float = float(amount_str)
                                # Handle both positive and negative amounts
                                is_positive_amount = amount_float > 0
                                amount_float = abs(amount_float)  # Make all amounts positive for display
                            except ValueError:
                                print(f"Skipping row {row_num} with invalid amount: {amount}")
                                continue
                            
                            # Always parse bank transaction for bank statements
                            if description:
                                parsed_category, merchant = self._parse_bank_transaction(description, amount_float, is_positive_amount)
                                category = parsed_category
                                # Use merchant name as description if it's cleaner
                                if len(merchant) < len(description):
                                    description = merchant
                                
                                # Check if this is a refund/rebate that should be negative
                                action_upper = description.upper()
                                is_refund_rebate = any(word in action_upper for word in ['REFUND', 'REBATE', 'RETURN', 'DEBIT CARD RETURN', 'ADJUST FEE', 'FEE REBATE'])
                                
                                # Make refunds/rebates negative amounts
                                if is_refund_rebate:
                                    amount_float = -amount_float
                            
                            # Don't skip any transactions - include all but mark non-spending as "No Category"
                            
                            # Format date properly
                            formatted_date = self._format_date(date) if date else datetime.date.today().isoformat()
                            
                            expense = {
                                "date": formatted_date,
                                "amount": amount_float,
                                "category": (category or 'other').lower(),
                                "description": description or '',
                                "notes": notes or ''  # Use notes from CSV if available, otherwise empty
                            }
                            self.expenses.append(expense)
                            loaded_count += 1
                            
                    except Exception as e:
                        print(f"Error processing row {row_num}: {e}")
                        continue
                        
            self.save_expenses()
            print(f"Successfully loaded {loaded_count} expenses from {csv_filename}")
            return loaded_count
            
        except FileNotFoundError:
            print(f"Error: CSV file '{csv_filename}' not found")
            return 0
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return 0

    def load_from_excel(self, excel_filename):
        """Load expenses from an Excel file"""
        if not EXCEL_AVAILABLE:
            print("❌ Excel support requires openpyxl. Installing now...")
            try:
                subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
                print("✅ openpyxl installed successfully! Please run the script again.")
                return 0
            except subprocess.CalledProcessError:
                print("❌ Failed to install openpyxl. Please install manually: pip install openpyxl")
                return 0
        
        # Clear existing expenses to avoid duplicates
        self.expenses = []
        loaded_count = 0
        
        try:
            from openpyxl import load_workbook
            
            # Load the workbook
            wb = load_workbook(excel_filename, read_only=True)
            
            # Look for the "All Transactions" sheet first, or use the first available sheet
            sheet = None
            if "All Transactions" in wb.sheetnames:
                sheet = wb["All Transactions"]
                print(f"Using 'All Transactions' sheet")
            else:
                # Find a sheet that looks like transaction data (has Date, Amount, Description columns)
                for sheet_name in wb.sheetnames:
                    test_sheet = wb[sheet_name]
                    headers = []
                    for cell in test_sheet[1]:
                        if cell.value:
                            headers.append(str(cell.value).lower().strip())
                        else:
                            headers.append('')
                    
                    # Check if this looks like transaction data
                    has_date = any('date' in h for h in headers)
                    has_amount = any(word in ' '.join(headers) for word in ['amount', 'transaction', 'debit', 'credit'])
                    
                    if has_date and has_amount:
                        sheet = test_sheet
                        print(f"Using sheet: '{sheet_name}' (detected transaction data)")
                        break
                
                # If no suitable sheet found, use the first one
                if sheet is None:
                    sheet = wb.worksheets[0]
                    print(f"Using first available sheet: '{sheet.title}'")
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).lower().strip())
                else:
                    headers.append('')
            
            print(f"Detected Excel columns: {headers}")
            
            # Process each row (skip header)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):  # Skip empty rows
                    continue
                    
                # Create a row dictionary
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and value is not None:
                        row_dict[headers[i]] = str(value) if value else ''
                
                # Skip rows that don't look like transactions
                if not row_dict:
                    continue
                
                # Parse the transaction - handle both raw bank data and already-categorized data
                try:
                    # Check if this is already categorized data (has category column)
                    if 'category' in row_dict and row_dict['category']:
                        # This is already processed data from our own export
                        amount_str = str(row_dict.get('amount', '0')).replace('$', '').replace(',', '')
                        try:
                            amount_float = float(amount_str)
                        except ValueError:
                            continue
                            
                        expense = {
                            "date": row_dict.get('date', ''),
                            "amount": amount_float,
                            "category": row_dict.get('category', 'other').lower(),
                            "description": row_dict.get('description', ''),
                            "notes": row_dict.get('notes', '')  # Preserve notes from Excel
                        }
                    else:
                        # This is raw bank data, needs parsing
                        amount = row_dict.get('amount ($)', '') or row_dict.get('amount', '')
                        description = row_dict.get('action', '') or row_dict.get('description', '')
                        date = row_dict.get('run date', '') or row_dict.get('date', '')
                        
                        if not amount or not description:
                            continue
                            
                        # Clean up amount
                        amount_str = str(amount).replace('$', '').replace(',', '')
                        try:
                            amount_float = float(amount_str)
                            is_positive_amount = amount_float > 0
                            amount_float = abs(amount_float)
                        except ValueError:
                            continue
                        
                        # Parse bank transaction with new logic
                        parsed_category, merchant = self._parse_bank_transaction(description, amount_float, is_positive_amount)
                        
                        # Check if this is a refund/rebate that should be negative
                        if description:
                            action_upper = description.upper()
                            is_refund_rebate = any(word in action_upper for word in ['REFUND', 'REBATE', 'RETURN', 'DEBIT CARD RETURN', 'ADJUST FEE', 'FEE REBATE'])
                            
                            # Make refunds/rebates negative amounts
                            if is_refund_rebate:
                                amount_float = -amount_float
                        
                        # Don't skip any transactions - include all
                            
                        expense = {
                            "date": date if date else datetime.date.today().isoformat(),
                            "amount": amount_float,
                            "category": parsed_category.lower(),
                            "description": merchant or description,
                            "notes": ''  # Initialize notes field for new transactions
                        }
                    
                    if expense:
                        self.expenses.append(expense)
                        loaded_count += 1
                except Exception as e:
                    print(f"Warning: Could not parse row {row_num}: {e}")
                    continue
            
            wb.close()
            self.save_expenses()
            print(f"Successfully loaded {loaded_count} expenses from {excel_filename}")
            return loaded_count
            
        except FileNotFoundError:
            print(f"Error: Excel file '{excel_filename}' not found")
            return 0
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return 0
    
    def _get_csv_value(self, row, possible_keys):
        """Helper to find value from CSV row using different possible column names"""
        for key in possible_keys:
            # Try exact match (case-insensitive)
            for row_key in row.keys():
                if row_key and row_key.lower().strip() == key.lower():
                    value = row[row_key]
                    if value is not None:
                        value = str(value).strip()
                        return value if value else None
                # Also try partial matches for compound column names
                if row_key and key.lower() in row_key.lower():
                    value = row[row_key]
                    if value is not None:
                        value = str(value).strip()
                        return value if value else None
        return None

    def _clean_csv_file(self, csv_filename):
        """Clean CSV file to handle BOM characters and empty lines"""
        import tempfile
        import os
        
        # Create a temporary cleaned file
        temp_fd, temp_filename = tempfile.mkstemp(suffix='.csv', text=True)
        
        try:
            with open(csv_filename, 'r', encoding='utf-8-sig') as input_file:
                with os.fdopen(temp_fd, 'w', encoding='utf-8') as output_file:
                    lines = input_file.readlines()
                    
                    # Find and write the header line
                    header_written = False
                    for line in lines:
                        stripped_line = line.strip()
                        if stripped_line and ('Run Date' in stripped_line or 'date' in stripped_line.lower() or 'amount' in stripped_line.lower()):
                            output_file.write(stripped_line + '\n')
                            header_written = True
                            break
                    
                    # Write all data lines after the header
                    if header_written:
                        reading_data = False
                        for line in lines:
                            stripped_line = line.strip()
                            # Skip empty lines
                            if not stripped_line:
                                continue
                            # Start reading data after header
                            if 'Run Date' in stripped_line or 'date' in stripped_line.lower():
                                reading_data = True
                                continue
                            # Write data lines
                            if reading_data:
                                output_file.write(stripped_line + '\n')
                    
            return temp_filename
            
        except Exception as e:
            # If cleaning fails, return original filename
            os.close(temp_fd)
            os.unlink(temp_filename)
            return csv_filename

    def _format_date(self, date_str):
        """Format date string to ISO format (YYYY-MM-DD)"""
        if not date_str:
            return datetime.date.today().isoformat()
        
        try:
            # Handle MM/DD/YYYY format
            if '/' in date_str:
                parts = date_str.split('/')
                if len(parts) == 3:
                    month, day, year = parts
                    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            
            # Handle other formats - add more as needed
            return date_str
            
        except Exception:
            return datetime.date.today().isoformat()

    def show_summary(self):
        """Display a summary of all expenses"""
        if not self.expenses:
            print("No expenses to summarize.")
            return
        
        from collections import defaultdict
        
        spending_totals = defaultdict(float)
        no_category_total = 0.0
        no_category_count = 0
        deposits_transfers_total = 0.0
        deposits_transfers_count = 0
        
        for expense in self.expenses:
            if expense['category'].lower() == 'no category':
                no_category_total += expense['amount']
                no_category_count += 1
                spending_totals[expense['category']] += expense['amount']  # Include No Category in spending
            elif expense['category'].lower() == 'deposits/transfers':
                deposits_transfers_total += expense['amount']
                deposits_transfers_count += 1
                # Don't include deposits/transfers in spending analysis
            else:
                spending_totals[expense['category']] += expense['amount']
        
        # Sort by amount
        sorted_spending = sorted(spending_totals.items(), key=lambda x: x[1], reverse=True)
        total_spending = sum(spending_totals.values())
        total_all = total_spending + deposits_transfers_total
        
        print(f"\n=== EXPENSE SUMMARY ===")
        print(f"Total expenses: {len(self.expenses)}")
        print(f"Total amount: ${total_all:,.2f}")
        
        print(f"\nBy category:")
        
        for category, amount in sorted_spending:
            if total_spending > 0:
                percentage = (amount / total_spending) * 100
                print(f"  {category.title()}: ${amount:,.2f} ({percentage:.1f}%)")
            else:
                print(f"  {category.title()}: ${amount:,.2f}")
        
        # Show summary of spending vs non-spending
        if deposits_transfers_count > 0:
            print(f"\n📊 Spending Analysis Summary:")
            print(f"  Spending transactions: {len(self.expenses) - deposits_transfers_count}, ${total_spending:,.2f}")
            print(f"  Non-spending transactions (Deposits/Transfers): {deposits_transfers_count}, ${deposits_transfers_total:,.2f}")
            print(f"  Total transactions: {len(self.expenses)}, ${total_all:,.2f}")

    def export_to_csv(self, output_filename=None):
        """Export expenses to a clean CSV file"""
        if not self.expenses:
            print("No expenses to export.")
            return None
        
        if not output_filename:
            # Generate filename based on current JSON filename
            base_name = self.filename.replace('.json', '')
            output_filename = f"{base_name}_categorized.csv"
        
        try:
            with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['date', 'amount', 'category', 'description', 'notes']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                for expense in sorted(self.expenses, key=lambda x: x['date']):
                    writer.writerow({
                        'date': expense['date'],
                        'amount': f"{expense['amount']:.2f}",
                        'category': expense['category'].title(),
                        'description': expense['description'],
                        'notes': expense.get('notes', '')
                    })
            
            print(f"📁 Exported {len(self.expenses)} expenses to: {output_filename}")
            return output_filename
            
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return None

    def _parse_bank_transaction(self, action_text, amount=0, is_positive_amount=False):
        """Parse bank transaction to extract merchant and category with smart rules"""
        action = action_text.upper()
        
        # Handle ATM transactions as financial
        if any(word in action for word in ['ATM', 'CASH ADVANCE']):
            return 'financial', action_text
        
        # Handle refunds and rebates - keep in appropriate category but make amount negative
        is_refund_rebate = any(word in action for word in ['REFUND', 'REBATE', 'RETURN', 'DEBIT CARD RETURN', 'ADJUST FEE', 'FEE REBATE'])
        
        # Mark deposits/credits and transfers as "Deposits/Transfers" (but not refunds/rebates)
        if not is_refund_rebate and any(word in action for word in ['CHECK RECEIVED', 'DEPOSIT', 'CREDIT', 'TRANSFER FROM', 'TRANSFER TO', 'TRANSFERRED FROM', 'TRANSFERRED TO', 'SAZERAC']):
            return 'deposits/transfers', action_text
        
        # Mark direct debits that are transfers/payments between accounts as "No Category" (but not refunds/rebates)
        if not is_refund_rebate and any(word in action for word in ['DIRECT DEBIT VENMO', 'DIRECT DEBIT PAYPAL', 'DIRECT DEBIT ZELLE', 'VENMO', 'ZELLE', 'PAYPAL']):
            return 'no category', action_text
        
        # Mark positive amounts (money coming in) as "Deposits/Transfers" unless it's a return/refund
        if is_positive_amount and not is_refund_rebate:
            # This is income/deposit, mark as "Deposits/Transfers"
            return 'deposits/transfers', action_text
        
        # Extract merchant from common patterns
        if 'DEBIT CARD PURCHASE' in action:
            # Look for merchant name after date pattern
            # Find merchant name between date and location/reference numbers
            match = re.search(r'\d{2}/\d{2} (.+?) \d{12}', action)
            if match:
                merchant = match.group(1).strip()
            else:
                # Fallback: extract text after "OUTSTAND AUTH MM/DD "
                match = re.search(r'OUTSTAND AUTH \d{2}/\d{2} (.+?) \d{12}', action)
                if match:
                    merchant = match.group(1).strip()
                else:
                    merchant = action_text
        elif 'DIRECT DEBIT' in action:
            # Extract company name from direct debit
            match = re.search(r'DIRECT DEBIT (.+?) PREAUTHPMT', action)
            if match:
                merchant = match.group(1).strip()
            else:
                merchant = action_text
        else:
            merchant = action_text
        
        # Get category using enhanced categorization
        category = self._categorize_merchant(merchant, amount)
        
        return category, merchant

    def _replace_account_numbers(self, description):
        """Replace account numbers with friendly names for deposits/transfers"""
        account_replacements = {
            'VS Z38-188212-1': 'Tax/Savings',
            'VS Z35-496083-1': 'Maddox Horse',
            'VS X91-527132-1': 'Pay Yo Bills',
            'VS X91-526994-1': '529 Holding Account/Savings',
            'VS Z25-701080-1': 'Maddox Checking',
            'VS Z27-576183-1': 'Maddox UTMA',
            'VS 618-173081-1': 'Maddox 529',
            'VS 618-453141-1': 'Coen 529',
            'VS Z25-686751-1': 'Coen UTMA',
            'VS 603-894285-1': 'Cece 529',
            'VS Z27-577046-1': 'Cece UTMA'            
        }
        
        # Replace account numbers with friendly names
        friendly_description = description
        for account_num, friendly_name in account_replacements.items():
            if account_num in friendly_description:
                friendly_description = friendly_description.replace(account_num, friendly_name)
        
        return friendly_description

    def _categorize_merchant(self, merchant, amount):
        """Enhanced merchant categorization with smart rules and vendor database"""
        merchant_lower = merchant.lower()
        
        # Define gas stations for amount-based categorization
        gas_stations = [
            'quiktrip', 'qt ', 'caseys', 'casey\'s', 'shell', 'exxon', 'mobil', 'chevron', 
            'bp', 'conoco', 'phillips 66', 'texaco', 'valero', 'citgo', '7-eleven', 
            'circle k', 'wawa', 'sheetz', 'loves', 'pilot', 'flying j', 'ta travel', 
            'speedway', 'marathon', 'sunoco', 'gulf', 'cenex'
        ]
        
        # Gas station logic: <$50 = misc, ≥$50 = transportation, BUT "OUTSIDE" = always transportation
        if any(station in merchant_lower for station in gas_stations):
            # Special case: QT/Quiktrip with "OUTSIDE" is always transportation regardless of amount
            if ('qt' in merchant_lower or 'quiktrip' in merchant_lower) and 'outside' in merchant_lower:
                return 'transportation'
            elif amount < 50:
                return 'misc'
            else:
                return 'transportation'
        
        # Sonic is always food
        if 'sonic' in merchant_lower:
            return 'food'
        
        # Walmart is always groceries
        if any(walmart in merchant_lower for walmart in ['walmart', 'wal-mart', 'walmart.com']):
            return 'groceries'
        
        # Enhanced vendor database with common merchants
        vendor_categories = {
            # Fast Food & Restaurants
            'food': [
                'taco bell', 'chick-fil-a', 'chick fil a', 'mcdonalds', 'mcdonald\'s', 'burger king',
                'subway', 'kfc', 'pizza hut', 'dominos', 'papa johns', 'wendy\'s', 'wendys',
                'arbys', 'arby\'s', 'popeyes', 'chipotle', 'panera', 'starbucks', 'dunkin',
                'ihop', 'dennys', 'denny\'s', 'waffle house', 'cracker barrel', 'olive garden',
                'applebees', 'chilis', 'tgi fridays', 'outback', 'red lobster', 'panda express',
                'qdoba', 'five guys', 'in-n-out', 'whataburger', 'culvers', 'culver\'s',
                'jimmy johns', 'jimmy john\'s', 'jersey mikes', 'firehouse subs', 'blaze pizza',
                'cafe', 'restaurant', 'diner', 'grill', 'bistro', 'eatery', 'coffee', 'bakery'
            ],
            # Transportation (non-gas station)
            'transportation': [
                'uber', 'lyft', 'taxi', 'parking', 'toll', 'metro', 'bus', 'train', 'airline',
                'airport', 'rental car', 'hertz', 'enterprise', 'budget', 'avis', 'national'
            ],
            # Groceries & Food Shopping
            'groceries': [
                'kroger', 'safeway', 'publix', 'whole foods', 'trader joes', 'trader joe\'s',
                'aldi', 'food lion', 'giant', 'stop shop', 'harris teeter', 'wegmans',
                'grocery', 'supermarket', 'fresh market', 'food store', 'butcher', 'deli'
            ],
            # Retail & Shopping (excluding groceries/walmart)
            'shopping': [
                'target', 'costco', 'sams club', 'sam\'s club', 'bjs', 'amazon',
                'best buy', 'home depot', 'lowes', 'lowe\'s', 'menards', 'ace hardware',
                'cvs', 'walgreens', 'rite aid', 'dollar general', 'dollar tree', 'family dollar',
                'macys', 'macy\'s', 'nordstrom', 'jcpenney', 'kohl\'s', 'kohls', 'sears',
                'old navy', 'gap', 'banana republic', 'tj maxx', 'marshalls', 'ross',
                'bed bath beyond', 'bath body works', 'victoria secret', 'victoria\'s secret',
                'store', 'shop', 'market', 'general', 'supply', 'depot'
            ],
            # Healthcare & Medical
            'healthcare': [
                'cvs pharmacy', 'walgreens pharmacy', 'pharmacy', 'hospital', 'clinic',
                'medical', 'doctor', 'dentist', 'dental', 'optometry', 'vision', 'urgent care',
                'pediatric', 'therapy', 'physical therapy', 'chiropractic', 'veterinary', 'vet'
            ],
            # Entertainment & Recreation
            'entertainment': [
                'movie', 'cinema', 'theater', 'theatre', 'netflix', 'spotify', 'apple music',
                'disney', 'hulu', 'amazon prime', 'xbox', 'playstation', 'nintendo', 'steam',
                'game', 'arcade', 'bowling', 'golf', 'gym', 'fitness', 'spa', 'salon',
                'amusement', 'zoo', 'aquarium', 'museum', 'concert', 'ticket', 'event'
            ],
            # Utilities & Services
            'utilities': [
                'electric', 'electricity', 'gas company', 'water', 'sewer', 'trash', 'waste',
                'internet', 'cable', 'phone', 'cellular', 'verizon', 'att', 'at&t', 't-mobile',
                'sprint', 'comcast', 'xfinity', 'spectrum', 'cox', 'dish', 'directv'
            ],
            # Financial & Professional Services
            'financial': [
                'bank', 'credit union', 'atm', 'paypal', 'venmo', 'zelle', 'cashapp', 'cash app',
                'square', 'stripe', 'insurance', ' tax ', 'accounting', 'legal', 'lawyer',
                'attorney', 'notary', 'real estate', 'mortgage', 'loan'
            ]
        }
        
        # Check vendor database
        for category, vendors in vendor_categories.items():
            if any(vendor in merchant_lower for vendor in vendors):
                return category
        
        # Pattern-based categorization for specific transaction types
        if any(word in merchant_lower for word in ['atm', 'cash advance', 'withdrawal']):
            return 'cash'
        elif any(word in merchant_lower for word in ['subscription', 'monthly', 'annual']):
            return 'subscription'
        elif 'sq *' in merchant_lower or 'square *' in merchant_lower:
            # Square payments - try to categorize by business name after SQ *
            sq_business = merchant_lower.replace('sq *', '').strip()
            # Recursively categorize the business name
            return self._categorize_merchant(sq_business, amount)
        
        # Default fallback
        return 'other'

    def display_summary(self):
        """Display a summary of all expenses"""
        if not self.expenses:
            print("No expenses recorded yet.")
            return
            
        total = sum(expense['amount'] for expense in self.expenses)
        categories = defaultdict(float)
        
        for expense in self.expenses:
            categories[expense['category']] += expense['amount']
        
        print(f"\n=== EXPENSE SUMMARY ===")
        print(f"Total expenses: {len(self.expenses)}")
        print(f"Total amount: ${total:.2f}")
        print(f"\nBy category:")
        for category, amount in sorted(categories.items()):
            print(f"  {category.title()}: ${amount:.2f}")
        
        print(f"\nAll expenses:")
        # Sort expenses by date ascending for display
        sorted_expenses = self._sort_expenses_by_date_ascending(self.expenses)
        for expense in sorted_expenses:
            print(f"  {expense['date']} | ${expense['amount']:.2f} | {expense['category']} | {expense['description']}")

    def _sort_expenses_by_date_ascending(self, expenses):
        """Sort expenses by date in ascending order (earliest first), then by amount descending"""
        from datetime import datetime
        
        def sort_key(expense):
            try:
                # Parse MM/DD/YYYY format
                parsed_date = datetime.strptime(expense['date'], '%m/%d/%Y')
            except ValueError:
                try:
                    # Try YYYY-MM-DD format as fallback
                    parsed_date = datetime.strptime(expense['date'], '%Y-%m-%d')
                except ValueError:
                    # If parsing fails, use the string date (fallback)
                    parsed_date = expense['date']
            return (parsed_date, -expense['amount'])  # Negative amount for descending order within same date
        
        return sorted(expenses, key=sort_key)

    def export_to_excel(self, output_filename=None, input_filename=None):
        """Export categorized expenses to an Excel file with multiple tabs and charts"""
        if not self.expenses:
            print("No expenses to export.")
            return None
        
        if not EXCEL_AVAILABLE:
            print("❌ Excel export requires openpyxl. Installing now...")
            try:
                subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
                print("✅ openpyxl installed successfully! Please run the script again.")
                return None
            except subprocess.CalledProcessError:
                print("❌ Failed to install openpyxl. Please install manually: pip install openpyxl")
                return None
        
        if output_filename is None:
            # Get current date for timestamp
            from datetime import datetime
            current_date = datetime.now().strftime('%Y-%m-%d')
            
            if input_filename:
                # Generate filename based on input filename with current date
                import os
                base_name = os.path.splitext(os.path.basename(input_filename))[0]
                
                # Check if this is a re-processing of a categorized file
                if "_categorized_" in base_name.lower():
                    # Extract month from the base name (assume it starts with month name)
                    month_name = base_name.split('_')[0] if '_' in base_name else base_name
                    output_filename = f"{month_name}_categorized_final_{current_date}.xlsx"
                else:
                    # First time processing
                    output_filename = f"{base_name}_categorized_{current_date}.xlsx"
            else:
                # Fallback to date range if no input filename provided
                dates = [expense['date'] for expense in self.expenses]
                
                # Parse dates properly for correct min/max calculation
                parsed_dates = []
                for date_str in dates:
                    try:
                        parsed_date = datetime.strptime(date_str, '%m/%d/%Y')
                        parsed_dates.append(parsed_date)
                    except ValueError:
                        try:
                            parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
                            parsed_dates.append(parsed_date)
                        except ValueError:
                            continue
                
                if parsed_dates:
                    min_date = min(parsed_dates).strftime('%m-%d-%Y')
                    max_date = max(parsed_dates).strftime('%m-%d-%Y')
                else:
                    min_date = min(dates).replace('/', '-')
                    max_date = max(dates).replace('/', '-')
                    
                output_filename = f"categorized_expenses_{min_date}_to_{max_date}_{current_date}.xlsx"
        
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create chart sheet FIRST (excluding "Deposits/Transfers")
            chart_sheet = wb.create_sheet("Category Chart")
            
            # We'll populate the chart after we process all the categories
            
            # Create summary sheet with all transactions SECOND
            summary_sheet = wb.create_sheet("All Transactions")
            
            # Add headers with styling
            headers = ['Date', 'Amount', 'Category', 'Description', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            # Sort and add all transactions by date ascending
            sorted_expenses = self._sort_expenses_by_date_ascending(self.expenses)
            for row, expense in enumerate(sorted_expenses, 2):
                # Apply account number replacements for deposits/transfers
                description = expense['description']
                if expense['category'].lower() == 'deposits/transfers':
                    description = self._replace_account_numbers(description)
                
                summary_sheet.cell(row=row, column=1, value=expense['date'])
                summary_sheet.cell(row=row, column=2, value=expense['amount'])
                summary_sheet.cell(row=row, column=3, value=expense['category'].title())
                summary_sheet.cell(row=row, column=4, value=description)
                summary_sheet.cell(row=row, column=5, value=expense.get('notes', ''))  # Use actual notes data
            
            # Auto-adjust column widths
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create "Needs Review" tab if this is a _final_ file (RIGHT AFTER All Transactions)
            if '_categorized_' in input_filename and 'final' in output_filename:
                needs_review_expenses = []
                
                # Add transactions with meaningful notes (not just spaces)
                for expense in self.expenses:
                    notes = expense.get('notes', '').strip()
                    if notes:  # Has non-empty notes after stripping spaces
                        needs_review_expenses.append(expense)
                
                # Add all "No Category" transactions
                for expense in self.expenses:
                    if expense['category'].lower() == 'no category':
                        # Avoid duplicates (in case a No Category transaction also has notes)
                        if expense not in needs_review_expenses:
                            needs_review_expenses.append(expense)
                
                if needs_review_expenses:
                    # Sort by notes column (transactions with notes first, then by notes content)
                    needs_review_expenses.sort(key=lambda x: (x.get('notes', '').strip() == '', x.get('notes', '').strip().lower()))
                    
                    review_sheet = wb.create_sheet("Needs Review")
                    
                    # Add headers with styling
                    headers = ['Date', 'Amount', 'Category', 'Description', 'Notes']
                    for col, header in enumerate(headers, 1):
                        cell = review_sheet.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Add data rows
                    for row, expense in enumerate(needs_review_expenses, 2):
                        # Apply account number replacements for deposits/transfers
                        description = expense['description']
                        if expense['category'].lower() == 'deposits/transfers':
                            description = self._replace_account_numbers(description)
                        
                        review_sheet.cell(row=row, column=1, value=expense['date'])
                        review_sheet.cell(row=row, column=2, value=expense['amount'])
                        review_sheet.cell(row=row, column=3, value=expense['category'].title())
                        review_sheet.cell(row=row, column=4, value=description)
                        review_sheet.cell(row=row, column=5, value=expense.get('notes', ''))
                    
                    # Auto-adjust column widths for review sheet
                    review_sheet.column_dimensions['A'].width = 12
                    review_sheet.column_dimensions['B'].width = 12
                    review_sheet.column_dimensions['C'].width = 18
                    review_sheet.column_dimensions['D'].width = 40
                    review_sheet.column_dimensions['E'].width = 30
                    
                    print(f"📝 Created 'Needs Review' tab with {len(needs_review_expenses)} transactions requiring attention")
            
            # Group expenses by category to get existing categories
            categories = defaultdict(list)
            spending_categories = defaultdict(list)  # Only for spending analysis
            
            for expense in self.expenses:
                categories[expense['category']].append(expense)
                # Only include in spending analysis if not "Deposits/Transfers"
                if expense['category'].lower() != 'deposits/transfers':
                    spending_categories[expense['category']].append(expense)
            
            # Add additional empty categories that we want tabs for
            additional_categories = ['horse', 'birthday/christmas', 'deposits/transfers', 'no category']
            for add_cat in additional_categories:
                if add_cat not in categories:
                    categories[add_cat] = []
            
            # Create a sheet for each category with dynamic formulas (including empty ones)
            # Sort categories by number of transactions (descending), with deposits/transfers always last
            def sort_categories_by_transactions(category_item):
                category, expenses = category_item
                if category.lower() == 'deposits/transfers':
                    return (0, len(expenses))  # Always last (0 priority)
                else:
                    return (1, len(expenses))  # Higher priority, sort by transaction count descending
            
            sorted_categories = sorted(categories.items(), key=sort_categories_by_transactions, reverse=True)
            
            for category, expenses in sorted_categories:
                # Sanitize sheet name for Excel (remove invalid characters)
                safe_sheet_name = category.title().replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
                sheet = wb.create_sheet(safe_sheet_name)
                
                # Add headers with styling for category sheets
                category_headers = ['Date', 'Amount', 'Category', 'Description', 'Notes']
                for col, header in enumerate(category_headers, 1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center")
                
                # Add simple formula-based approach for dynamic updates
                # Use basic Excel functions that work across versions
                max_row = len(self.expenses) + 1  # +1 for header row
                
                # Add a simple instruction row for this category
                sheet.cell(row=2, column=1, value=f"=IF(COUNTIF('All Transactions'!$C$2:$C${max_row},\"{category.title()}\")>0,\"Category data below\",\"No {category.title()} transactions\")")
                
                # Add the current matching transactions as static data
                matching_expenses = [expense for expense in self.expenses if expense['category'] == category]
                
                # Add the current matching transactions starting from row 3
                for row_num, expense in enumerate(matching_expenses, 3):
                    # Apply account number replacements for deposits/transfers
                    description = expense['description']
                    if category.lower() == 'deposits/transfers':
                        description = self._replace_account_numbers(description)
                    
                    sheet.cell(row=row_num, column=1, value=expense['date'])
                    sheet.cell(row=row_num, column=2, value=expense['amount'])
                    sheet.cell(row=row_num, column=3, value=expense['category'].title())
                    sheet.cell(row=row_num, column=4, value=description)
                    sheet.cell(row=row_num, column=5, value=expense.get('notes', ''))  # Use actual notes data
                
                # Special handling for Deposits/Transfers tab - add Cash In/Cash Out summary
                if category.lower() == 'deposits/transfers' and matching_expenses:
                    self._add_cash_flow_summary(sheet, matching_expenses, len(matching_expenses) + 5)
                
                # Add instructions for empty categories
                if not matching_expenses:
                    sheet.cell(row=3, column=1, value="To add transactions to this category:")
                    sheet.cell(row=4, column=1, value="1. Go to 'All Transactions' tab")
                    sheet.cell(row=5, column=1, value="2. Change any transaction's Category column")
                    sheet.cell(row=6, column=1, value="3. Re-run the script to update")
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Now populate the chart sheet with data (after all categories are processed)
            # Get unique spending categories for dynamic chart (exclude only "Deposits/Transfers")
            unique_spending_categories = [cat for cat in categories.keys() if cat.lower() != 'deposits/transfers']
            max_row = len(self.expenses) + 1  # +1 for header row
            
            # Add chart data headers
            chart_sheet['A1'] = 'Category'
            chart_sheet['B1'] = 'Total'
            chart_sheet['A1'].font = Font(bold=True)
            chart_sheet['B1'].font = Font(bold=True)
            
            # Add dynamic SUMIF formulas for chart totals (widely compatible, excluding "Deposits/Transfers")
            chart_row = 2
            print(f"📊 Debug: Creating chart with categories: {sorted(unique_spending_categories)}")
            for category in sorted(unique_spending_categories):
                # Category name (use sanitized name for display)
                display_name = category.title().replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
                chart_sheet.cell(row=chart_row, column=1, value=display_name)
                
                # Use SUMIF formula which is widely supported across Excel versions
                # Use the original category name (not title case) to match the actual data
                sum_formula = f'=SUMIF(\'All Transactions\'!$C:$C,"{category}",\'All Transactions\'!$B:$B)'
                chart_sheet.cell(row=chart_row, column=2, value=sum_formula)
                
                print(f"📊 Debug: Added chart row {chart_row}: {display_name} with formula: {sum_formula}")
                chart_row += 1
            
            # Add total row at the bottom
            chart_sheet.cell(row=chart_row, column=1, value="TOTAL").font = Font(bold=True)
            total_formula = f'=SUM(B2:B{chart_row-1})'
            total_cell = chart_sheet.cell(row=chart_row, column=2, value=total_formula)
            total_cell.font = Font(bold=True)
            
            # Add separator line above total
            chart_sheet.cell(row=chart_row-1, column=1).border = None  # This will be handled by Excel formatting
            
            # Create pie chart (exclude the total row)
            pie = PieChart()
            # Include all category rows but exclude the total row (chart_row-1 is the last category row)
            labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=chart_row-1)  # Include all categories, exclude total
            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=chart_row-1)    # Include all categories, exclude total
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Expenses by Category"
            pie.height = 15
            pie.width = 20
            
            # Add data labels
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
            pie.dataLabels.showVal = True
            
            # Add chart to sheet
            chart_sheet.add_chart(pie, "D3")
            
            # Auto-adjust column widths for chart sheet
            chart_sheet.column_dimensions['A'].width = 20
            chart_sheet.column_dimensions['B'].width = 15
            
            # Save the workbook
            wb.save(output_filename)
            
            print(f"\n✅ Exported {len(self.expenses)} categorized expenses to: {output_filename}")
            print(f"📊 Created {len(categories)} category tabs plus summary and chart tabs")
            
            # Show summary with spending vs non-spending breakdown
            spending_count = sum(len(expenses) for cat, expenses in categories.items() if cat.lower() != 'deposits/transfers')
            non_spending_count = len(categories.get('deposits/transfers', []))
            spending_total = sum(sum(exp['amount'] for exp in expenses) for cat, expenses in categories.items() if cat.lower() != 'deposits/transfers')
            non_spending_total = sum(exp['amount'] for exp in categories.get('deposits/transfers', []))
            
            print(f"\nExported categories:")
            for category in sorted(categories.keys()):
                count = len(categories[category])
                total = sum(expense['amount'] for expense in categories[category])
                if category.lower() == 'deposits/transfers':
                    print(f"  {category.title()}: {count} transactions, ${total:.2f} (non-spending)")
                else:
                    print(f"  {category.title()}: {count} transactions, ${total:.2f}")
            
            if non_spending_count > 0:
                print(f"\n📊 Spending Analysis Summary:")
                print(f"  Spending transactions: {spending_count}, ${spending_total:.2f}")
                print(f"  Non-spending transactions: {non_spending_count}, ${non_spending_total:.2f}")
                print(f"  Total transactions: {len(self.expenses)}, ${spending_total + non_spending_total:.2f}")
            
            # Auto-download the file
            self._download_to_computer(output_filename)
            
            return output_filename
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return None

    def _download_to_computer(self, filename):
        """Download the CSV file to the user's local computer"""
        try:
            # Get the absolute path to the file
            file_path = os.path.abspath(filename)
            
            # Check if file exists
            if not os.path.exists(file_path):
                print(f"❌ File not found: {file_path}")
                return
            
            print(f"\n📥 Downloading {filename} to your computer...")
            
            # Use VS Code's built-in download functionality
            # This works in development containers and codespaces
            download_command = f'code --download "{file_path}"'
            
            try:
                # Try VS Code download first
                result = subprocess.run(download_command, shell=True, capture_output=True, text=True, timeout=10)
                if result.returncode == 0:
                    print(f"✅ Successfully downloaded {filename} to your Downloads folder!")
                    print(f"💡 Check your browser's Downloads folder for the file.")
                    return
            except (subprocess.TimeoutExpired, subprocess.CalledProcessError, FileNotFoundError):
                pass
            
            # Fallback: Provide instructions for manual download
            print(f"📋 Manual download instructions:")
            print(f"   1. In VS Code, look at the left sidebar (File Explorer)")
            print(f"   2. Right-click on '{filename}'")
            print(f"   3. Select 'Download' from the context menu")
            print(f"   4. The file will be saved to your Downloads folder")
            print(f"   5. Alternative: Click the file to open it, then Ctrl+S to save")
            
            # Also try using the browser download method if in a web environment
            if os.environ.get('CODESPACE_NAME') or os.environ.get('GITPOD_WORKSPACE_ID'):
                print(f"\n🌐 Alternative: Open the file in browser:")
                print(f"   File location: /workspaces/Expenses_Scripts/{filename}")
                print(f"   Or use: File menu → Download → navigate to {filename}")
                
        except Exception as e:
            print(f"❌ Error during download: {e}")
            print(f"📁 File saved in workspace at: /workspaces/Expenses_Scripts/{filename}")
            print(f"💡 You can manually download it using VS Code's file explorer.")

    def _add_cash_flow_summary(self, sheet, deposits_transfers_expenses, start_row):
        """Add Cash In/Cash Out summary table to Deposits/Transfers sheet"""
        try:
            # Categories for Cash In vs Cash Out
            cash_in_keywords = ['transferred from', 'direct deposit', 'check received', 'deposit', 'credit']
            cash_out_keywords = ['transferred to', 'direct debit', 'payment', 'transfer to']
            
            # Track cash flows by account/description
            cash_in_accounts = defaultdict(float)
            cash_out_accounts = defaultdict(float)
            
            for expense in deposits_transfers_expenses:
                description = self._replace_account_numbers(expense['description']).lower()
                amount = abs(expense['amount'])  # Use absolute value for calculations
                
                # Determine if this is cash in or cash out
                is_cash_in = any(keyword in description for keyword in cash_in_keywords)
                is_cash_out = any(keyword in description for keyword in cash_out_keywords)
                
                if is_cash_in:
                    # Extract account name for cash in
                    if 'transferred from' in description:
                        account_match = description.split('transferred from')[-1].strip()
                        account_name = account_match.split(' ')[0:3]  # Take first few words
                        account_name = ' '.join(account_name).title()
                    elif 'direct deposit' in description:
                        if 'sazerac' in description:
                            account_name = 'Sazerac Payroll'
                        else:
                            account_name = 'Direct Deposit'
                    elif 'check received' in description:
                        account_name = 'Check Received'
                    else:
                        account_name = 'Other Cash In'
                    
                    cash_in_accounts[account_name] += amount
                    
                elif is_cash_out:
                    # Extract account name for cash out
                    if 'transferred to' in description:
                        account_match = description.split('transferred to')[-1].strip()
                        account_name = account_match.split(' ')[0:3]  # Take first few words
                        account_name = ' '.join(account_name).title()
                    elif 'direct debit' in description:
                        if 'venmo' in description:
                            account_name = 'Venmo'
                        elif 'paypal' in description:
                            account_name = 'PayPal'
                        else:
                            account_name = 'Direct Debit'
                    else:
                        account_name = 'Other Cash Out'
                    
                    cash_out_accounts[account_name] += amount
                
                # If neither, try to categorize by amount sign
                elif expense['amount'] > 0:
                    cash_in_accounts['Unclassified Cash In'] += amount
                else:
                    cash_out_accounts['Unclassified Cash Out'] += amount
            
            # Add the summary table
            current_row = start_row
            
            # Title
            title_cell = sheet.cell(row=current_row, column=1, value="💰 CASH FLOW SUMMARY")
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            sheet.merge_cells(f'A{current_row}:E{current_row}')
            current_row += 2
            
            # Cash In section
            cash_in_header = sheet.cell(row=current_row, column=1, value="💵 CASH IN")
            cash_in_header.font = Font(bold=True, size=12, color="FFFFFF")
            cash_in_header.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            amount_header = sheet.cell(row=current_row, column=2, value="Amount")
            amount_header.font = Font(bold=True, size=12, color="FFFFFF")
            amount_header.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            current_row += 1
            
            cash_in_total = 0
            for account, amount in sorted(cash_in_accounts.items()):
                sheet.cell(row=current_row, column=1, value=account)
                sheet.cell(row=current_row, column=2, value=amount)
                cash_in_total += amount
                current_row += 1
            
            # Cash In total
            total_cell = sheet.cell(row=current_row, column=1, value="TOTAL CASH IN")
            total_cell.font = Font(bold=True)
            total_amount_cell = sheet.cell(row=current_row, column=2, value=cash_in_total)
            total_amount_cell.font = Font(bold=True)
            current_row += 2
            
            # Cash Out section
            cash_out_header = sheet.cell(row=current_row, column=1, value="💸 CASH OUT")
            cash_out_header.font = Font(bold=True, size=12, color="FFFFFF")
            cash_out_header.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            
            amount_header = sheet.cell(row=current_row, column=2, value="Amount")
            amount_header.font = Font(bold=True, size=12, color="FFFFFF")
            amount_header.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            current_row += 1
            
            cash_out_total = 0
            for account, amount in sorted(cash_out_accounts.items()):
                sheet.cell(row=current_row, column=1, value=account)
                sheet.cell(row=current_row, column=2, value=amount)
                cash_out_total += amount
                current_row += 1
            
            # Cash Out total
            total_cell = sheet.cell(row=current_row, column=1, value="TOTAL CASH OUT")
            total_cell.font = Font(bold=True)
            total_amount_cell = sheet.cell(row=current_row, column=2, value=cash_out_total)
            total_amount_cell.font = Font(bold=True)
            current_row += 2
            
            # Net Cash Flow
            net_flow = cash_in_total - cash_out_total
            net_cell = sheet.cell(row=current_row, column=1, value="NET CASH FLOW")
            net_cell.font = Font(bold=True, size=12)
            net_amount_cell = sheet.cell(row=current_row, column=2, value=net_flow)
            net_amount_cell.font = Font(bold=True, size=12)
            
            # Color the net flow based on positive/negative
            if net_flow >= 0:
                net_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                net_amount_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            else:
                net_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                net_amount_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            # Auto-adjust column widths for summary
            sheet.column_dimensions['A'].width = max(25, sheet.column_dimensions['A'].width or 0)
            sheet.column_dimensions['B'].width = max(15, sheet.column_dimensions['B'].width or 0)
            
            print(f"💰 Added Cash Flow Summary: ${cash_in_total:.2f} in, ${cash_out_total:.2f} out, Net: ${net_flow:.2f}")
            
        except Exception as e:
            print(f"Warning: Could not create cash flow summary: {e}")
  
# Main execution
if __name__ == "__main__":  
    tracker = ExpenseTracker()
    
    # Check if file provided as command line argument
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        
        # Detect file type based on extension
        file_extension = input_file.lower().split('.')[-1]
        
        if file_extension in ['xlsx', 'xls']:
            print(f"Loading expenses from Excel file: {input_file}")
            loaded = tracker.load_from_excel(input_file)
        elif file_extension == 'csv':
            print(f"Loading expenses from CSV file: {input_file}")
            loaded = tracker.load_from_csv(input_file)
        else:
            print(f"❌ Unsupported file type: .{file_extension}")
            print("📝 Supported file types: .csv, .xlsx, .xls")
            loaded = 0
        
        if loaded > 0:
            tracker.display_summary()
            
            # Automatically export to Excel with multiple tabs and charts
            exported_file = tracker.export_to_excel(input_filename=input_file)
            if exported_file:
                print(f"\n🎉 Ready to use! Your categorized expenses are now in: {exported_file}")
                print(f"📊 Multiple tabs available: All Transactions, individual category tabs, and Category Chart")
                print(f"💡 You can open this file in Excel, Google Sheets, or any spreadsheet program.")
        else:
            print("No expenses were loaded from the input file.")
    else:
        # Original test behavior - add sample expenses
        print("No input file provided. Running with test data...")
        print("Usage: python expenses.py <filename>")
        print("Example: python expenses.py my_expenses.csv")
        print("Example: python expenses.py my_expenses.xlsx")
        print("📝 Supported file types: .csv, .xlsx, .xls\n")
        
        # Add a few test expenses  
        tracker.add_expense(12.50, "Food", "Lunch at cafe")  
        tracker.add_expense(45.00, "Transportation", "Gas")  
        tracker.add_expense(8.99, "Food", "Coffee")  
          
        # Show what we have so far  
        tracker.display_summary()
        
        # Export test data 
        tracker.export_to_excel()  