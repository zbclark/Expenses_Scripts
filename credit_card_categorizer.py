#!/usr/bin/env python3
"""
Credit Card Transaction Categorizer
Categorizes credit card transactions into Work, Personal, or Budgeted categories
"""

import csv
import os
import sys
import glob
from datetime import datetime
import re
from collections import defaultdict
import subprocess

# Check for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class CreditCardCategorizer:
    def __init__(self):
        self.budgeted_vendors = {
            # Subscription services with specific amounts where noted
            'xbox': 'Budgeted',
            'microsoft': 'Budgeted', 
            'scoop it dude': 'Budgeted',
            'aa wifi': 'Budgeted',
            'naked wines': 'Budgeted',
            'nakedwines.com': 'Budgeted',
            'emerge': 'Budgeted',
            'kindle': 'Budgeted',  # (11.99)
            'appleone': 'Budgeted',  # (37.95)
            'spotify': 'Budgeted',
            'youtubetv': 'Budgeted',
            'youtube tv': 'Budgeted',
            'netflix': 'Budgeted',
            'apple': 'Budgeted',  # (3.99, 0.99, and 37.95) - but needs careful matching
            'apple.com': 'Budgeted',  # (11.99 and 39.99) - specific amount matching
            'hulu': 'Budgeted',
            'google': 'Budgeted',  # (9.99 and 19.99) - needs careful matching
            'actblue': 'Budgeted',
            'nyt crossword': 'Budgeted',
            'ny times': 'Budgeted',
            'new york times': 'Budgeted',
            'grint': 'Budgeted',
            'audible': 'Budgeted',  # (149.50)
            'carrot': 'Budgeted',
            'paramount+': 'Budgeted',
            'paramount plus': 'Budgeted',
            'disney+': 'Budgeted',
            'disney plus': 'Budgeted',
            'dropbox': 'Budgeted',
            'sams club': 'Budgeted',
            'sam\'s club': 'Budgeted',
            'prime video': 'Budgeted',  # (4.99)
            'rob israel': 'Budgeted',
            'jenks public works': 'Budgeted',
            'municipal online': 'Budgeted',  # (1.25)
            }
        
        # Work-related vendor patterns
        self.work_vendors = {
            'aws': 'Work',
            'amazon web services': 'Work',
            'github': 'Work',
            'zoom': 'Work',
            'office': 'Work',
            'adobe': 'Work',
            'jetbrains': 'Work',
            'atlassian': 'Work',
            'slack': 'Work',
            'notion': 'Work',
            'figma': 'Work',
            'canva': 'Work',
            'hotel': 'Work',
            'airline': 'Work',
            'lyft': 'Work',  # Could be work travel
            'rental car': 'Work',
            'hertz': 'Work',
            'enterprise': 'Work',
            'national': 'Work',  # National car rental
            'conference': 'Work',
            'seminar': 'Work',
            'training': 'Work'
        }
        
        # Birthday/Christmas vendor patterns
        self.birthday_christmas_vendors = {
            'etsy': 'Birthday/Christmas',
            'amazon': 'Birthday/Christmas',  # Could be gifts, needs context
            'target': 'Birthday/Christmas',  # Could be gifts, needs context
            'walmart': 'Birthday/Christmas',  # Could be gifts, needs context
            'toys r us': 'Birthday/Christmas',
            'toysrus': 'Birthday/Christmas',
            'hallmark': 'Birthday/Christmas',
            'party city': 'Birthday/Christmas',
            'gift': 'Birthday/Christmas',
            'present': 'Birthday/Christmas',
            'birthday': 'Birthday/Christmas',
            'christmas': 'Birthday/Christmas',
            'holiday': 'Birthday/Christmas',
            'wrapping': 'Birthday/Christmas',
            'card': 'Birthday/Christmas',  # Greeting cards
            'balloon': 'Birthday/Christmas',
            'decoration': 'Birthday/Christmas',
            'cake': 'Birthday/Christmas',
            'bakery': 'Birthday/Christmas',
            'flower': 'Birthday/Christmas',
            'florist': 'Birthday/Christmas'
        }
        
        # Horse-related vendor patterns
        self.horse_vendors = {
            'tack': 'Horse',
            'equestrian': 'Horse',
            'horse': 'Horse',
            'stable': 'Horse',
            'horse barn': 'Horse',
            'boarding barn': 'Horse',
            'feed barn': 'Horse',
            'feed': 'Horse',
            'hay': 'Horse',
            'veterinary': 'Horse',
            'vet': 'Horse',
            'farrier': 'Horse',
            'saddle': 'Horse',
            'bridle': 'Horse',
            'halter': 'Horse',
            'grooming': 'Horse',
            'supplement': 'Horse',
            'blanket': 'Horse',
            'boot': 'Horse',  # Horse boots
            'lesson': 'Horse',
            'training': 'Horse',  # Could conflict with work, needs priority
            'riding': 'Horse',
            'arena': 'Horse',
            'competition': 'Horse',
            'show': 'Horse',  # Horse shows
            'farm house tack': 'Horse',
            'showgrounds': 'Horse',
            'ridervideo': 'Horse',
            'pine ridge equine': 'Horse'
        }
        
        self.transactions = []
    
    def detect_csv_format(self, filepath):
        """Detect the format/structure of the CSV file"""
        with open(filepath, 'r', encoding='utf-8-sig') as file:
            # Read first few lines to detect format
            lines = []
            for i, line in enumerate(file):
                lines.append(line.strip())
                if i >= 5:  # Read first 6 lines
                    break
        
        # Look for common credit card CSV headers
        header_patterns = {
            'chase': ['transaction date', 'description', 'amount'],
            'amex': ['date', 'description', 'amount'],
            'discover': ['trans. date', 'description', 'amount'],
            'citi': ['date', 'description', 'debit', 'credit'],
            'capital_one': ['transaction date', 'description', 'debit', 'credit'],
            'generic': ['date', 'description', 'amount']
        }
        
        # Convert lines to lowercase for matching
        content_lower = '\n'.join(lines).lower()
        
        # Detect format based on headers
        for format_name, keywords in header_patterns.items():
            if all(keyword in content_lower for keyword in keywords):
                print(f"📊 Detected format: {format_name.title()}")
                return format_name
        
        print("📊 Using generic format")
        return 'generic'
    
    def parse_csv_file(self, filepath):
        """Parse a single CSV file and extract transactions"""
        print(f"\n📄 Processing: {os.path.basename(filepath)}")
        
        format_type = self.detect_csv_format(filepath)
        transactions = []
        
        try:
            with open(filepath, 'r', encoding='utf-8-sig') as file:
                # Skip any empty lines at the beginning
                content = file.read().strip()
                lines = content.split('\n')
                
                # Find the header line (look for common patterns)
                header_line_idx = 0
                for i, line in enumerate(lines):
                    line_lower = line.lower()
                    if any(word in line_lower for word in ['date', 'description', 'amount', 'trans']):
                        header_line_idx = i
                        break
                
                # Parse from header line
                csv_content = '\n'.join(lines[header_line_idx:])
                reader = csv.DictReader(csv_content.splitlines())
                
                for row_num, row in enumerate(reader, 1):
                    try:
                        transaction = self.extract_transaction_data(row, format_type, filepath)
                        if transaction:
                            transaction['source_file'] = os.path.basename(filepath)
                            transaction['card_account'] = self.extract_card_account(filepath)
                            transaction['purchaser'] = self.determine_purchaser(filepath, row)
                            transactions.append(transaction)
                    except Exception as e:
                        print(f"   Warning: Could not parse row {row_num}: {e}")
                        continue
                
        except Exception as e:
            print(f"   Error reading file: {e}")
            return []
        
        print(f"   ✅ Loaded {len(transactions)} transactions")
        return transactions
    
    def extract_card_account(self, filepath):
        """Extract card account identifier from filename"""
        filename = os.path.basename(filepath)
        # Look for pattern like x0111, x3823, etc.
        match = re.search(r'x(\d+)', filename.lower())
        if match:
            return f"Card x{match.group(1)}"
        return "Unknown Card"
    
    def extract_transaction_data(self, row, format_type, filepath=None):
        """Extract transaction data based on detected format"""
        # Common field name variations
        date_fields = ['transaction date', 'trans. date', 'date', 'post date', 'posting date']
        desc_fields = ['description', 'desc', 'merchant', 'reference']
        amount_fields = ['amount', 'debit', 'credit']
        
        # Find the actual field names (case-insensitive)
        row_keys = {k.lower(): k for k in row.keys()}
        
        # Extract date
        date_val = None
        for field in date_fields:
            if field in row_keys:
                date_val = row[row_keys[field]]
                break
        
        # Extract description
        desc_val = None
        for field in desc_fields:
            if field in row_keys:
                desc_val = row[row_keys[field]]
                if desc_val and desc_val.strip():
                    break
        
        # Extract amount (handle debit/credit columns)
        amount_val = 0.0
        if 'debit' in row_keys and 'credit' in row_keys:
            # Handle separate debit/credit columns
            debit = row[row_keys['debit']].strip() if row[row_keys['debit']] else ''
            credit = row[row_keys['credit']].strip() if row[row_keys['credit']] else ''
            
            if debit and debit != '0':
                amount_val = -abs(float(debit.replace('$', '').replace(',', '')))
            elif credit and credit != '0':
                amount_val = abs(float(credit.replace('$', '').replace(',', '')))
        else:
            # Handle single amount column
            for field in amount_fields:
                if field in row_keys:
                    amount_str = row[row_keys[field]].strip()
                    if amount_str:
                        # Clean and convert amount
                        amount_str = amount_str.replace('$', '').replace(',', '')
                        amount_val = float(amount_str)
                        break
        
        # Skip if missing essential data
        if not date_val or not desc_val or amount_val == 0:
            return None
        
        # Skip payment transactions (card payments, transfers, etc.)
        desc_lower = desc_val.lower().strip()
        payment_keywords = [
            'payment', 'autopay', 'automatic payment', 'online payment', 
            'thank you', 'transfer', 'bill pay', 'payoff', 'balance transfer',
            'credit card payment', 'payment received', 'payment - thank you'
        ]
        
        if any(keyword in desc_lower for keyword in payment_keywords):
            return None
        
        # Format date
        formatted_date = self.format_date(date_val)
        
        return {
            'date': formatted_date,
            'description': desc_val.strip(),
            'amount': amount_val,
            'raw_row': dict(row)
        }
    
    def determine_purchaser(self, filepath, row):
        """Determine who made the purchase based on file and member name"""
        filename = os.path.basename(filepath).lower()
        
        # Extract card number from filename
        if 'x9449' in filename or 'x8379' in filename:
            return 'Zac'
        elif 'x0111' in filename or 'x7336' in filename:
            return 'Jenny'
        elif 'x3823' in filename:
            # For x3823, use Member Name column
            member_name = row.get('Member Name', '') or row.get('member name', '')
            if member_name and member_name.strip():
                member_name = member_name.strip().upper()
                if 'ZAC CLARK' in member_name or 'ZACHARY CLARK' in member_name:
                    return 'Zac'
                elif 'JENNIFER' in member_name or 'CLARK' in member_name:
                    return 'Jenny'
            # Default to Jenny for x3823 
            return 'Jenny'
        
        # Default fallback
        return 'Unknown'
    
    def create_transaction_key(self, transaction):
        """Create a unique key for a transaction to match against existing categorizations"""
        # Use date, amount, and first 50 chars of description as unique identifier
        return f"{transaction['date']}|{transaction['amount']:.2f}|{transaction['description'][:50]}"
    
    def load_existing_categorizations(self):
        """Load existing categorizations from previous Excel export"""
        existing = {}
        
        # Look for existing Excel file with today's date
        timestamp = datetime.now().strftime('%Y-%m-%d')
        excel_filename = f"credit_card_categorized_{timestamp}.xlsx"
        
        if not os.path.exists(excel_filename):
            print(f"📄 No existing categorized file found ({excel_filename})")
            return existing
        
        print(f"📄 Loading existing categorizations from {excel_filename}...")
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_filename)
            
            # Read from "All Transactions" sheet
            if "All Transactions" in wb.sheetnames:
                sheet = wb["All Transactions"]
                
                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append(str(cell.value).lower().strip())
                    else:
                        headers.append('')
                
                # Find column indices
                date_col = headers.index('date') if 'date' in headers else 0
                amount_col = headers.index('amount') if 'amount' in headers else 3
                desc_col = headers.index('description') if 'description' in headers else 2
                category_col = headers.index('category') if 'category' in headers else 4
                notes_col = headers.index('notes') if 'notes' in headers else 7
                
                # Process each row (skip header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row) or len(row) <= max(date_col, amount_col, desc_col):
                        continue
                    
                    try:
                        date = str(row[date_col]) if row[date_col] else ''
                        amount = float(row[amount_col]) if row[amount_col] else 0
                        description = str(row[desc_col]) if row[desc_col] else ''
                        category = str(row[category_col]) if len(row) > category_col and row[category_col] else ''
                        notes = str(row[notes_col]) if len(row) > notes_col and row[notes_col] else ''
                        
                        # Create transaction key
                        key = f"{date}|{amount:.2f}|{description[:50]}"
                        existing[key] = {
                            'category': category,
                            'notes': notes
                        }
                    except (ValueError, IndexError):
                        continue
                        
                print(f"✅ Loaded {len(existing)} existing categorizations")
            
        except Exception as e:
            print(f"⚠️  Could not load existing categorizations: {e}")
        
        return existing
    
    def format_date(self, date_str):
        """Format date string to consistent format"""
        if not date_str:
            return ''
        
        date_str = date_str.strip()
        
        # Try common date formats
        formats = ['%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y', '%d/%m/%Y']
        
        for fmt in formats:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                return date_obj.strftime('%Y-%m-%d')
            except ValueError:
                continue
        
        # Return original if can't parse
        return date_str
    
    def categorize_transaction(self, transaction):
        """Categorize a single transaction into Work, Personal, Budgeted, Birthday/Christmas, or Horse"""
        description = transaction['description'].lower()
        amount = abs(transaction['amount'])
        
        # Check budgeted vendors first (highest priority for subscriptions)
        for vendor, category in self.budgeted_vendors.items():
            if vendor in description:
                # Special handling for vendors with specific amounts
                if vendor == 'apple' and category == 'Budgeted':
                    # Only categorize as budgeted if amount matches expected subscription amounts
                    if abs(amount - 3.99) < 0.01 or abs(amount - 0.99) < 0.01 or abs(amount - 37.95) < 0.01:
                        return 'Budgeted'
                elif vendor == 'google' and category == 'Budgeted':
                    # Only categorize as budgeted if amount matches subscription amounts
                    if abs(amount - 9.99) < 0.01 or abs(amount - 19.99) < 0.01:
                        return 'Budgeted'
                elif vendor == 'kindle' and category == 'Budgeted':
                    # Check for Kindle Unlimited subscription amount
                    if abs(amount - 11.99) < 0.01:
                        return 'Budgeted'
                elif vendor == 'audible' and category == 'Budgeted':
                    # Check for Audible annual subscription
                    if abs(amount - 149.50) < 0.01 or abs(amount - 14.95) < 0.01:
                        return 'Budgeted'
                elif vendor == 'municipal online' and category == 'Budgeted':
                    # Check for specific municipal payment amount
                    if abs(amount - 1.25) < 0.01:
                        return 'Budgeted'
                elif vendor == 'prime video' and category == 'Budgeted':
                    # Check for Prime Video subscription amount
                    if abs(amount - 4.99) < 0.01:
                        return 'Budgeted'
                elif vendor == 'apple.com' and category == 'Budgeted':
                    # Check for Apple.com subscription amounts
                    if abs(amount - 11.99) < 0.01 or abs(amount - 39.99) < 0.01 or abs(amount - 37.95) < 0.01:
                        return 'Budgeted'
                else:
                    return category
        
        # Check horse vendors (high priority - specific category)
        for vendor, category in self.horse_vendors.items():
            if vendor in description:
                # Handle conflicts: 'training' could be work or horse - use context
                if vendor == 'training':
                    # If it contains horse-related terms, categorize as horse
                    if any(horse_term in description for horse_term in ['horse', 'riding', 'equestrian', 'lesson']):
                        return 'Horse'
                    # Otherwise let it fall through to work check
                    continue
                return category
        
        # Check birthday/christmas vendors (medium priority)
        for vendor, category in self.birthday_christmas_vendors.items():
            if vendor in description:
                # Handle broad vendors that could be multiple categories
                if vendor in ['amazon', 'target', 'walmart']:
                    # Only categorize as Birthday/Christmas if description has gift indicators
                    gift_indicators = ['gift', 'present', 'birthday', 'christmas', 'holiday', 'party']
                    if any(indicator in description for indicator in gift_indicators):
                        return 'Birthday/Christmas'
                    # Otherwise continue to other checks
                    continue
                elif vendor == 'card':
                    # Only categorize as Birthday/Christmas if it's clearly a greeting card
                    if any(term in description for term in ['hallmark', 'greeting', 'birthday', 'christmas']):
                        return 'Birthday/Christmas'
                    # Otherwise continue (could be credit card, etc.)
                    continue
                else:
                    return category
        
        # Check work vendors (lower priority than specific categories)
        for vendor, category in self.work_vendors.items():
            if vendor in description:
                return category
        
        # Default to Personal
        return 'Personal'
    
    def process_all_csv_files(self):
        """Process CSV files and auto-categorize transactions"""
        print("📂 Processing CSV files in current directory...")
        return self.process_csv_files()
    
    def load_transactions_from_excel(self, excel_filename):
        """Load transactions from existing Excel file"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_filename)
            
            # Read from "All Transactions" sheet
            if "All Transactions" not in wb.sheetnames:
                print("❌ No 'All Transactions' sheet found in Excel file")
                return self.process_csv_files()
            
            sheet = wb["All Transactions"]
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).lower().strip())
                else:
                    headers.append('')
            
            print(f"📊 Detected Excel columns: {headers}")
            
            # Find column indices
            date_col = headers.index('date') if 'date' in headers else 0
            card_col = headers.index('card account') if 'card account' in headers else 1
            desc_col = headers.index('description') if 'description' in headers else 2
            amount_col = headers.index('amount') if 'amount' in headers else 3
            category_col = headers.index('category') if 'category' in headers else 4
            purchaser_col = headers.index('purchaser') if 'purchaser' in headers else 5
            source_col = headers.index('source file') if 'source file' in headers else 6
            notes_col = headers.index('notes') if 'notes' in headers else 7
            
            all_transactions = []
            
            # Process each row (skip header)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):  # Skip empty rows
                    continue
                
                try:
                    date = str(row[date_col]) if len(row) > date_col and row[date_col] else ''
                    card_account = str(row[card_col]) if len(row) > card_col and row[card_col] else ''
                    description = str(row[desc_col]) if len(row) > desc_col and row[desc_col] else ''
                    amount_str = str(row[amount_col]) if len(row) > amount_col and row[amount_col] else '0'
                    existing_category = str(row[category_col]) if len(row) > category_col and row[category_col] else ''
                    purchaser = str(row[purchaser_col]) if len(row) > purchaser_col and row[purchaser_col] else ''
                    source_file = str(row[source_col]) if len(row) > source_col and row[source_col] else ''
                    notes = str(row[notes_col]) if len(row) > notes_col and row[notes_col] else ''
                    
                    # Parse amount
                    amount_str = amount_str.replace('$', '').replace(',', '')
                    try:
                        amount = float(amount_str)
                    except ValueError:
                        continue
                    
                    transaction = {
                        'date': date,
                        'card_account': card_account,
                        'description': description,
                        'amount': amount,
                        'category': existing_category,  # Keep existing category for now
                        'purchaser': purchaser,
                        'source_file': source_file,
                        'notes': notes
                    }
                    
                    all_transactions.append(transaction)
                    
                except (ValueError, IndexError) as e:
                    print(f"   Warning: Could not parse row {row_num}: {e}")
                    continue
            
            if not all_transactions:
                print("❌ No valid transactions found in Excel file")
                return self.process_csv_files()
            
            print(f"✅ Loaded {len(all_transactions)} transactions from Excel file")
            
            # Keep existing categorizations and notes - do NOT re-categorize
            print(f"\n📊 Using existing categorizations from Excel file...")
            
            categorized_stats = defaultdict(int)
            
            for transaction in all_transactions:
                # Keep the existing category from Excel - do NOT re-categorize
                categorized_stats[transaction['category']] += 1
            
            self.transactions = all_transactions
            
            # Show summary
            print(f"\n📊 Categorization Summary:")
            for category, count in sorted(categorized_stats.items()):
                total_amount = sum(abs(t['amount']) for t in all_transactions if t['category'] == category)
                print(f"   {category}: {count} transactions, ${total_amount:,.2f}")
            
            return all_transactions
            
        except Exception as e:
            print(f"❌ Error loading from Excel file: {e}")
            print("📂 Falling back to CSV processing...")
            return self.process_csv_files()
    
    def process_csv_files(self):
        """Process all CSV files in the current directory (fallback method)"""
        csv_files = glob.glob('*.csv') + glob.glob('*.CSV')
        
        if not csv_files:
            print("❌ No CSV files found in current directory")
            return
        
        print(f"🔍 Found {len(csv_files)} CSV files to process:")
        for file in csv_files:
            print(f"   - {file}")
        
        all_transactions = []
        
        for csv_file in csv_files:
            transactions = self.parse_csv_file(csv_file)
            all_transactions.extend(transactions)
        
        if not all_transactions:
            print("❌ No transactions found in any CSV files")
            return
        
        # Filter out "Unknown Card" transactions (keep only specific card transactions)
        original_count = len(all_transactions)
        all_transactions = [t for t in all_transactions if t['card_account'] != 'Unknown Card']
        filtered_count = len(all_transactions)
        
        if original_count != filtered_count:
            print(f"🧹 Filtered out {original_count - filtered_count} duplicate 'Unknown Card' transactions")
        
        # Categorize all transactions
        print(f"\n🏷️  Categorizing {filtered_count} transactions...")
        
        categorized_stats = defaultdict(int)
        
        for transaction in all_transactions:
            transaction['category'] = self.categorize_transaction(transaction)
            transaction['notes'] = ''  # Empty notes for new transactions
            categorized_stats[transaction['category']] += 1
        
        self.transactions = all_transactions
        
        # Show summary
        print(f"\n📊 Categorization Summary:")
        for category, count in sorted(categorized_stats.items()):
            total_amount = sum(abs(t['amount']) for t in all_transactions if t['category'] == category)
            print(f"   {category}: {count} transactions, ${total_amount:,.2f}")
        
        return all_transactions
    
    def create_transaction_key(self, transaction):
        """Create a unique key for a transaction to match against existing categorizations"""
        # Use date, amount, and first 50 chars of description as unique identifier
        return f"{transaction['date']}|{transaction['amount']:.2f}|{transaction['description'][:50]}"
    
    def export_categorized_transactions(self, output_filename=None, export_format='both', input_filename=None):
        """Export categorized transactions to CSV and/or Excel"""
        if not self.transactions:
            print("❌ No transactions to export")
            return None
        
        timestamp = datetime.now().strftime('%Y-%m-%d')
        
        # Generate filename based on input filename (like expenses.py does)
        if input_filename and not output_filename:
            base_name = os.path.splitext(os.path.basename(input_filename))[0]
            
            # Check if this is a re-processing of a categorized file
            if "_categorized_" in base_name.lower():
                # Extract month or base name
                if '_' in base_name:
                    parts = base_name.split('_')
                    if len(parts) >= 3:  # credit_card_categorized_2024-11-03
                        base_part = '_'.join(parts[:2])  # credit_card_categorized
                    else:
                        base_part = base_name
                else:
                    base_part = base_name
                output_base = f"{base_part}_final_{timestamp}"
            else:
                # First time processing
                output_base = f"{base_name}_categorized_{timestamp}"
        else:
            # Fallback to default naming
            output_base = f"credit_card_categorized_{timestamp}"
        
        # Export to CSV
        csv_filename = None
        if export_format in ['csv', 'both']:
            csv_filename = output_filename or f"{output_base}.csv"
            self._export_to_csv(csv_filename)
        
        # Export to Excel
        excel_filename = None
        if export_format in ['excel', 'both']:
            if output_filename and output_filename.endswith('.csv'):
                excel_filename = output_filename.replace('.csv', '.xlsx')
            else:
                excel_filename = f"{output_base}.xlsx"
            
            if EXCEL_AVAILABLE:
                self._export_to_excel(excel_filename)
            else:
                print("❌ Excel export requires openpyxl. Installing now...")
                try:
                    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
                    print("✅ openpyxl installed successfully! Please run the script again for Excel export.")
                except subprocess.CalledProcessError:
                    print("❌ Failed to install openpyxl. Please install manually: pip install openpyxl")
        
        return excel_filename or csv_filename
    
    def _export_to_csv(self, filename):
        """Export transactions to CSV format"""
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['date', 'card_account', 'description', 'amount', 'category', 'source_file']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                
                # Sort by date and amount
                sorted_transactions = sorted(self.transactions, key=lambda x: (x['date'], -abs(x['amount'])))
                
                for transaction in sorted_transactions:
                    writer.writerow({
                        'date': transaction['date'],
                        'card_account': transaction['card_account'],
                        'description': transaction['description'],
                        'amount': f"{transaction['amount']:.2f}",
                        'category': transaction['category'],
                        'source_file': transaction['source_file']
                    })
            
            print(f"📄 CSV exported to: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Error exporting to CSV: {e}")
            return None
    
    def _export_to_excel(self, filename):
        """Export transactions to Excel with multiple tabs and charts"""
        try:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create summary sheet with all transactions FIRST
            summary_sheet = wb.create_sheet("All Transactions")
            
            # Add headers with styling
            headers = ['Date', 'Card Account', 'Description', 'Amount', 'Category', 'Purchaser', 'Source File', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            # Sort and add all transactions
            sorted_transactions = sorted(self.transactions, key=lambda x: (x['date'], -abs(x['amount'])))
            for row, transaction in enumerate(sorted_transactions, 2):
                summary_sheet.cell(row=row, column=1, value=transaction['date'])
                summary_sheet.cell(row=row, column=2, value=transaction['card_account'])
                summary_sheet.cell(row=row, column=3, value=transaction['description'])
                summary_sheet.cell(row=row, column=4, value=transaction['amount'])
                summary_sheet.cell(row=row, column=5, value=transaction['category'])
                summary_sheet.cell(row=row, column=6, value=transaction.get('purchaser', 'Unknown'))
                summary_sheet.cell(row=row, column=7, value=transaction['source_file'])
                summary_sheet.cell(row=row, column=8, value=transaction.get('notes', ''))  # Notes column
            
            # Auto-adjust column widths
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Group transactions by category
            categories = defaultdict(list)
            for transaction in self.transactions:
                categories[transaction['category']].append(transaction)
            
            # Add all expected categories (even if empty) like the main expenses script
            all_categories = ['Work', 'Personal', 'Budgeted', 'Birthday/Christmas', 'Horse']
            for category in all_categories:
                if category not in categories:
                    categories[category] = []
            
            # Create individual category sheets (not "X Transactions" but just the category name)
            # Sort categories by predefined order, then by transaction count
            category_order = ['Work', 'Personal', 'Budgeted', 'Birthday/Christmas', 'Horse']
            sorted_categories = []
            for cat in category_order:
                if cat in categories:
                    sorted_categories.append((cat, categories[cat]))
            
            # Add any other categories found that weren't in our predefined list
            for cat, transactions in categories.items():
                if cat not in category_order:
                    sorted_categories.append((cat, transactions))
            
            for category, transactions in sorted_categories:
                # Sanitize sheet name for Excel (Birthday/Christmas -> Birthday-Christmas)
                safe_sheet_name = category.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
                sheet = wb.create_sheet(safe_sheet_name)
                
                # Add headers with category-specific styling
                category_colors = {
                    'Work': "4CAF50",            # Green
                    'Personal': "2196F3",        # Blue  
                    'Budgeted': "FF9800",        # Orange
                    'Birthday-Christmas': "E91E63",  # Pink (note the dash)
                    'Horse': "795548"            # Brown
                }
                color = category_colors.get(safe_sheet_name, "366092")
                
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Add category transactions or placeholder if empty
                if transactions:
                    # Special sorting for Personal category - group by purchaser
                    if category == 'Personal':
                        # Group transactions by purchaser
                        purchaser_groups = defaultdict(list)
                        for transaction in transactions:
                            purchaser = transaction.get('purchaser', 'Unknown')
                            purchaser_groups[purchaser].append(transaction)
                        
                        current_row = 2
                        for purchaser in sorted(purchaser_groups.keys()):
                            purchaser_transactions = purchaser_groups[purchaser]
                            
                            # Add purchaser header
                            if len(purchaser_groups) > 1:  # Only add header if there are multiple purchasers
                                sheet.cell(row=current_row, column=1, value=f"{purchaser} Transactions:").font = Font(bold=True, size=12)
                                current_row += 1
                            
                            # Sort purchaser's transactions by date
                            sorted_purchaser_transactions = sorted(purchaser_transactions, key=lambda x: (x['date'], -abs(x['amount'])))
                            
                            # Add transactions for this purchaser
                            for transaction in sorted_purchaser_transactions:
                                sheet.cell(row=current_row, column=1, value=transaction['date'])
                                sheet.cell(row=current_row, column=2, value=transaction['card_account'])
                                sheet.cell(row=current_row, column=3, value=transaction['description'])
                                sheet.cell(row=current_row, column=4, value=transaction['amount'])
                                sheet.cell(row=current_row, column=5, value=transaction['category'])
                                sheet.cell(row=current_row, column=6, value=transaction.get('purchaser', 'Unknown'))
                                sheet.cell(row=current_row, column=7, value=transaction['source_file'])
                                sheet.cell(row=current_row, column=8, value=transaction.get('notes', ''))  # Notes column
                                current_row += 1
                            
                            # Add 2 rows of space between purchasers (except after the last one)
                            if purchaser != list(sorted(purchaser_groups.keys()))[-1]:
                                current_row += 2
                    else:
                        # Default sorting for all other categories
                        sorted_cat_transactions = sorted(transactions, key=lambda x: (x['date'], -abs(x['amount'])))
                        for row, transaction in enumerate(sorted_cat_transactions, 2):
                            sheet.cell(row=row, column=1, value=transaction['date'])
                            sheet.cell(row=row, column=2, value=transaction['card_account'])
                            sheet.cell(row=row, column=3, value=transaction['description'])
                            sheet.cell(row=row, column=4, value=transaction['amount'])
                            sheet.cell(row=row, column=5, value=transaction['category'])
                            sheet.cell(row=row, column=6, value=transaction.get('purchaser', 'Unknown'))
                            sheet.cell(row=row, column=7, value=transaction['source_file'])
                            sheet.cell(row=row, column=8, value=transaction.get('notes', ''))  # Notes column
                    
                    # Add category summary in column K (col 11), starting at row 2 (only once per sheet)
                    summary_col = 11  # Column K
                    summary_row = 2
                    total_amount = sum(abs(t['amount']) for t in transactions)
                    sheet.cell(row=summary_row, column=summary_col, value=f"{category} Summary:").font = Font(bold=True, size=14)
                    sheet.cell(row=summary_row + 1, column=summary_col, value=f"Total Transactions: {len(transactions)}")
                    sheet.cell(row=summary_row + 2, column=summary_col, value=f"Total Amount: ${total_amount:,.2f}")
                else:
                    # Add placeholder for empty categories like the main expenses script
                    sheet.cell(row=2, column=1, value=f"No {category} transactions found")
                    sheet.cell(row=3, column=1, value="Transactions will appear here when categorized")
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # (Removed duplicate summary block at the bottom of the sheet)
            
            # Create chart sheet
            chart_sheet = wb.create_sheet("Category Summary")
            
            # Add chart data headers
            chart_sheet['A1'] = 'Category'
            chart_sheet['B1'] = 'Count'
            chart_sheet['C1'] = 'Total Amount'
            
            # Style headers
            for cell in ['A1', 'B1', 'C1']:
                chart_sheet[cell].font = Font(bold=True)
                chart_sheet[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                chart_sheet[cell].font = Font(bold=True, color="FFFFFF")
            
            # Add category data
            chart_row = 2
            for category, transactions in sorted_categories:
                total_amount = sum(abs(t['amount']) for t in transactions)
                chart_sheet.cell(row=chart_row, column=1, value=category)
                chart_sheet.cell(row=chart_row, column=2, value=len(transactions))
                chart_sheet.cell(row=chart_row, column=3, value=total_amount)
                chart_row += 1
            
            # Add purchaser breakdown for Personal and Work categories
            breakdown_start_row = chart_row + 3
            
            # Personal category breakdown
            personal_transactions = [t for t in self.transactions if t['category'] == 'Personal']
            if personal_transactions:
                chart_sheet.cell(row=breakdown_start_row, column=1, value='Personal Category Breakdown by Purchaser').font = Font(bold=True, size=14)
                chart_sheet.cell(row=breakdown_start_row + 1, column=1, value='Purchaser').font = Font(bold=True)
                chart_sheet.cell(row=breakdown_start_row + 1, column=2, value='Count').font = Font(bold=True)
                chart_sheet.cell(row=breakdown_start_row + 1, column=3, value='Total Amount').font = Font(bold=True)
                
                # Group personal transactions by purchaser
                personal_by_purchaser = defaultdict(list)
                for t in personal_transactions:
                    personal_by_purchaser[t.get('purchaser', 'Unknown')].append(t)
                
                breakdown_row = breakdown_start_row + 2
                for purchaser, transactions in sorted(personal_by_purchaser.items()):
                    total_amount = sum(abs(t['amount']) for t in transactions)
                    chart_sheet.cell(row=breakdown_row, column=1, value=purchaser)
                    chart_sheet.cell(row=breakdown_row, column=2, value=len(transactions))
                    chart_sheet.cell(row=breakdown_row, column=3, value=total_amount)
                    breakdown_row += 1
                
                breakdown_start_row = breakdown_row + 2
            
            # Work category breakdown
            work_transactions = [t for t in self.transactions if t['category'] == 'Work']
            if work_transactions:
                chart_sheet.cell(row=breakdown_start_row, column=1, value='Work Category Breakdown by Purchaser').font = Font(bold=True, size=14)
                chart_sheet.cell(row=breakdown_start_row + 1, column=1, value='Purchaser').font = Font(bold=True)
                chart_sheet.cell(row=breakdown_start_row + 1, column=2, value='Count').font = Font(bold=True)
                chart_sheet.cell(row=breakdown_start_row + 1, column=3, value='Total Amount').font = Font(bold=True)
                
                # Group work transactions by purchaser
                work_by_purchaser = defaultdict(list)
                for t in work_transactions:
                    work_by_purchaser[t.get('purchaser', 'Unknown')].append(t)
                
                breakdown_row = breakdown_start_row + 2
                for purchaser, transactions in sorted(work_by_purchaser.items()):
                    total_amount = sum(abs(t['amount']) for t in transactions)
                    chart_sheet.cell(row=breakdown_row, column=1, value=purchaser)
                    chart_sheet.cell(row=breakdown_row, column=2, value=len(transactions))
                    chart_sheet.cell(row=breakdown_row, column=3, value=total_amount)
                    breakdown_row += 1
            
            # Auto-adjust column widths for chart sheet
            chart_sheet.column_dimensions['A'].width = 15
            chart_sheet.column_dimensions['B'].width = 12
            chart_sheet.column_dimensions['C'].width = 15
            
            # Save workbook
            wb.save(filename)
            
            print(f"📊 Excel exported to: {filename}")
            print(f"   - All Transactions tab with {len(self.transactions)} total transactions")
            print(f"   - {len(categories)} individual category tabs")
            print(f"   - Category Summary tab with purchaser breakdowns")
            
            # Download the file
            self._download_to_computer(filename)
            
            return filename
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")
            return None
    
    def _download_to_computer(self, filename):
        """Download the file to the user's local computer"""
        try:
            # Get the absolute path to the file
            file_path = os.path.abspath(filename)
            
            # Check if file exists
            if not os.path.exists(file_path):
                print(f"❌ File not found: {file_path}")
                return
            
            print(f"\n📥 Downloading {filename} to your computer...")
            
            # Use VS Code's built-in download functionality
            download_command = f'code --download "{file_path}"'
            
            try:
                # Try VS Code download first
                result = subprocess.run(download_command, shell=True, capture_output=True, text=True, timeout=10)
                if result.returncode == 0:
                    print(f"✅ Successfully downloaded {filename} to your Downloads folder!")
                    print(f"💡 Check your browser's Downloads folder for the file.")
                    return
            except (subprocess.TimeoutExpired, subprocess.CalledProcessError, FileNotFoundError):
                pass
            
            # Fallback: Provide instructions for manual download
            print(f"📋 Manual download instructions:")
            print(f"   1. In VS Code, look at the left sidebar (File Explorer)")
            print(f"   2. Right-click on '{filename}'")
            print(f"   3. Select 'Download' from the context menu")
            print(f"   4. The file will be saved to your Downloads folder")
                
        except Exception as e:
            print(f"❌ Error during download: {e}")
            print(f"📁 File saved in workspace at: /workspaces/Expenses_Scripts/{filename}")
    
    def show_budgeted_vendor_matches(self):
        """Show which budgeted vendors were found in the transactions"""
        if not self.transactions:
            print("❌ No transactions loaded")
            return
        
        budgeted_transactions = [t for t in self.transactions if t['category'] == 'Budgeted']
        
        if not budgeted_transactions:
            print("📊 No budgeted transactions found")
            return
        
        print(f"\n💰 Found {len(budgeted_transactions)} Budgeted Transactions:")
        print("=" * 60)
        
        vendor_matches = defaultdict(list)
        
        for transaction in budgeted_transactions:
            description = transaction['description'].lower()
            
            # Find which vendor pattern matched
            matched_vendor = None
            for vendor in self.budgeted_vendors.keys():
                if vendor in description:
                    matched_vendor = vendor
                    break
            
            if matched_vendor:
                vendor_matches[matched_vendor].append(transaction)
        
        for vendor, transactions in sorted(vendor_matches.items()):
            total_amount = sum(abs(t['amount']) for t in transactions)
            print(f"\n🏷️  {vendor.title()}:")
            print(f"   Count: {len(transactions)} transactions")
            print(f"   Total: ${total_amount:.2f}")
            
            # Show individual transactions
            for t in sorted(transactions, key=lambda x: x['date']):
                print(f"   {t['date']} | {t['card_account']} | ${t['amount']:.2f} | {t['description'][:50]}")

def main():
    categorizer = CreditCardCategorizer()
    
    print("💳 Credit Card Transaction Categorizer")
    print("=" * 50)
    
    # Check if file provided as command line argument
    input_file = None
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        print(f"📄 Input file specified: {input_file}")
        
        # Check if this is a _categorized_ file
        if "_categorized_" in os.path.basename(input_file).lower():
            print("� Re-processing categorized file - preserving existing categorizations...")
            transactions = categorizer.load_transactions_from_excel(input_file)
        else:
            print("📂 Processing new file...")
            transactions = categorizer.process_all_csv_files()
    else:
        print("📂 Processing CSV files in current directory...")
        transactions = categorizer.process_all_csv_files()
    
    if not transactions:
        return
    
    # Export categorized results to Excel only
    print(f"\n📤 Exporting results...")
    output_file = categorizer.export_categorized_transactions(export_format='excel', input_filename=input_file)
    
    # Show detailed category breakdown
    categories = defaultdict(list)
    for t in transactions:
        categories[t['category']].append(t)
    
    print(f"\n📋 Detailed Category Breakdown:")
    for category in sorted(categories.keys()):
        cat_transactions = categories[category]
        count = len(cat_transactions)
        total = sum(abs(t['amount']) for t in cat_transactions)
        print(f"\n🏷️  {category}: {count} transactions, ${total:,.2f}")
        
        # Show top 5 vendors by amount for each category
        vendor_totals = defaultdict(float)
        for t in cat_transactions:
            # Extract vendor name (first few words)
            vendor = ' '.join(t['description'].split()[:3])
            vendor_totals[vendor] += abs(t['amount'])
        
        top_vendors = sorted(vendor_totals.items(), key=lambda x: x[1], reverse=True)[:5]
        for vendor, amount in top_vendors:
            print(f"     - {vendor}: ${amount:.2f}")
    
    # Show budgeted vendor analysis
    categorizer.show_budgeted_vendor_matches()
    
    if output_file:
        print(f"\n🎉 Processing complete!")
        print(f"� Excel file created with categorized transactions")
        print(f"📈 Excel file includes multiple tabs and charts for analysis")

if __name__ == "__main__":
    main()